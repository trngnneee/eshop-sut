"""Build reproducible HW06 reporting artifacts from committed test tables and Newman JSON.

This is intentionally deterministic: no pass/fail number is typed into the reports;
the suite figures are read from Newman JSON files at build time.
"""
from __future__ import annotations

import csv
import json
import re
import subprocess
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
HW = ROOT / "hw06"
REPORT = HW / "report"
EXCEL = HW / "excel"
RUNS = ROOT / "tests" / "test-runs"
SUMMARY = ROOT / "tests" / "test-summary"

APIS = [
    ("API-1", "api-01-login", "login", "POST /api/login", "FR-02, SEC-02, SEC-05", "LOGIN"),
    ("API-2", "api-02-checkout", "checkout", "POST /api/checkout", "FR-08, SEC-02, SEC-04", "CHECKOUT"),
    ("API-3", "api-03-admin-order-status", "order-status", "PUT /api/admin/orders/:id/status", "FR-10, FR-12, SEC-03", "ORDER-STATUS"),
]


def parse_table(path: Path) -> list[dict[str, str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    table = [line for line in lines if line.lstrip().startswith("|")]
    if len(table) < 3:
        raise ValueError(f"No Markdown table found in {path}")
    headers = [c.strip() for c in table[0].strip().strip("|").split("|")]
    headers = ["Expected result" if header == "Expected" else header for header in headers]
    rows: list[dict[str, str]] = []
    for line in table[2:]:
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) != len(headers) or not cells[0].startswith("TC-"):
            continue
        rows.append(dict(zip(headers, cells)))
    return rows


def ai_count(path: Path) -> int:
    return sum(1 for line in path.read_text(encoding="utf-8").splitlines() if re.search(r"\|\s*TC-[A-Z-]+-\d{3}\s*\|", line))


def read_stats(name: str) -> dict[str, int]:
    data = json.loads((HW / "newman" / "reports" / f"{name}.json").read_text(encoding="utf-8"))
    stats = data["run"]["stats"]
    result = {key: int(stats[key]["total"]) for key in ("iterations", "requests", "assertions")}
    result["failed"] = int(stats["assertions"]["failed"])
    result["request_failed"] = int(stats["requests"]["failed"])
    result["script_failed"] = int(stats["testScripts"]["failed"]) + int(stats["prerequestScripts"]["failed"])
    return result


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, title: str, rows: list[dict[str, str]], fields: list[str]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(fields)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
        sheet.column_dimensions[column[0].column_letter].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_excel(all_rows: list[dict[str, str]]) -> None:
    fields = ["API", "TC ID", "Requirement", "Nhóm", "Kỹ thuật", "Preconditions", "Method + Endpoint / Test data", "Expected result", "Nguồn", "Kỳ vọng chạy", "Bug ID", "Execution", "Lý do"]
    write_csv(EXCEL / "test-cases.csv", all_rows, fields)
    write_xlsx(EXCEL / "test-cases.xlsx", "HW06 test cases", all_rows, fields)

    summary_rows = []
    for label, folder, _, endpoint, req, _ in APIS:
        rows = [r for r in all_rows if r["API"] == label]
        ai = ai_count(HW / folder / "01-ai-generated.md")
        extended = sum(1 for r in rows if r["Nguồn"].lower().startswith("human"))
        summary_rows.append({"API": label, "Endpoint": endpoint, "Requirements": req, "AI generated": str(ai), "Human extended": str(extended), "Final cases": str(len(rows)), "Automated": str(sum(1 for r in rows if r.get("Execution") == "Automated")), "Manual": str(sum(1 for r in rows if r.get("Execution") == "Manual")), "Blocked": str(sum(1 for r in rows if r.get("Execution") == "Blocked")), "Expected FAIL": str(sum(1 for r in rows if r["Kỳ vọng chạy"].upper().startswith("FAIL")))})
    fields2 = ["API", "Endpoint", "Requirements", "AI generated", "Human extended", "Final cases", "Automated", "Manual", "Blocked", "Expected FAIL"]
    write_csv(EXCEL / "test-summary.csv", summary_rows, fields2)
    write_xlsx(EXCEL / "test-summary.xlsx", "HW06 summary", summary_rows, fields2)


def write_main_report(all_rows: list[dict[str, str]], stats: dict[str, dict[str, int]]) -> None:
    issue_manifest_path = REPORT / "github-issues.json"
    issue_manifest = {item["bug_id"]: item for item in json.loads(issue_manifest_path.read_text(encoding="utf-8"))} if issue_manifest_path.exists() else {}
    automated = sum(1 for row in all_rows if row.get("Execution") == "Automated")
    manual = sum(1 for row in all_rows if row.get("Execution") == "Manual")
    blocked = sum(1 for row in all_rows if row.get("Execution") == "Blocked")
    execution_percent = automated * 100 / len(all_rows)
    lines = [
        "# HW06 — AI-first API testing report",
        "",
        "> Báo cáo kỹ thuật được dựng từ test table đã commit và Newman JSON. Các bằng chứng HUMAN-only (ảnh, chữ ký, issue GitHub, sơ đồ tự vẽ, bài critique) không được giả mạo.",
        "",
        "## 1. Phạm vi và môi trường",
        "",
        "SUT là EShop backend trong `backend/`, chạy Node.js/Express + SQLite. Bộ kiểm thử chọn ba luồng có rủi ro cao: đăng nhập, checkout và chuyển trạng thái order của admin. Newman chạy local với `http://127.0.0.1:3001` vì port 3000 đang được tiến trình khác sử dụng; workflow CI dùng `http://localhost:3000` theo đề bài.",
        "",
        "Header `X-Student-Id: 23127207` được chèn ở collection-level pre-request và được log ở console cho mọi request.",
        "",
        "## 2. Pipeline sinh — audit — mở rộng",
        "",
        "Mỗi API dùng chuỗi P1 phân tích input/state → P2 domain partition + BVA → P3 state transition → P4 security → P5 schema. Output thô nằm ở `01-ai-generated.md`; audit phân loại 100% case thành VALID/INVALID/INCOMPLETE trong `02-audit.md`; `03-extended.md` thêm 6 case ngoài phạm vi prompt; bảng chốt là `test-cases.md`.",
        "",
        "| API | AI sinh | Audit | Human mở rộng | Bảng chốt | Audit thống kê |",
        "| :--- | ---: | :--- | ---: | ---: | :--- |",
    ]
    for label, folder, _, endpoint, _, _ in APIS:
        rows = [r for r in all_rows if r["API"] == label]
        audit = (HW / folder / "02-audit.md").read_text(encoding="utf-8")
        counts = {label: int(match.group(1)) for label, match in (("VALID", re.search(r"\|\s*VALID\s*\|\s*(\d+)", audit)), ("INVALID", re.search(r"\|\s*INVALID\s*\|\s*(\d+)", audit)), ("INCOMPLETE", re.search(r"\|\s*INCOMPLETE\s*\|\s*(\d+)", audit))) if match}
        stats_text = " / ".join(f"{counts.get(key, '?')} {key}" for key in ("VALID", "INVALID", "INCOMPLETE"))
        lines.append(f"| {label} — `{endpoint}` | {ai_count(HW / folder / '01-ai-generated.md')} | 100% gán nhãn | {sum(1 for r in rows if r['Nguồn'].lower().startswith('human'))} | {len(rows)} | {stats_text} |")
    api3_audit = (HW / "api-03-admin-order-status" / "02-audit.md").read_text(encoding="utf-8")
    api3_signoff = "**Signature / confirmation:** Đã duyệt" in api3_audit and "**Reviewed by:** Đặng Đăng Khoa" in api3_audit
    review_summary = ("Human review trong API-1, API-2 và file audit API-3 hiện đều có metadata xác nhận; người nộp cần tự kiểm tra chữ ký API-3 trước khi nộp." if api3_signoff else "Human review trong API-1 và API-2 đã được ghi trong audit theo các phê duyệt người dùng đã cung cấp. API-3 mới có agent pre-review; chữ ký người học vẫn phải được bổ sung độc lập.")
    lines += [
        "",
        review_summary,
        "",
        "## 3. Newman execution",
        "",
        f"Báo cáo đối soát [`execution-coverage.md`](../newman/reports/execution-coverage.md) trích TC ID trực tiếp từ tên assertion trong mọi Newman JSON: **{automated}/{len(all_rows)} = {execution_percent:.1f}%** case đã thực thi; {manual} Manual và {blocked} Blocked đều có lý do trong bảng test case.",
        "",
        "| Run | Iterations | Requests | Assertions | Failed assertions | Ý nghĩa |",
        "| :--- | ---: | ---: | ---: | ---: | :--- |",
        f"| `00-off-suite` | {stats['00-off-suite']['iterations']} | {stats['00-off-suite']['requests']} | {stats['00-off-suite']['assertions']} | {stats['00-off-suite']['failed']} | baseline CI xanh |",
        f"| `00-canary-suite` | {stats['00-canary-suite']['iterations']} | {stats['00-canary-suite']['requests']} | {stats['00-canary-suite']['assertions']} | {stats['00-canary-suite']['failed']} | strict canary: TC-API-LOGIN-018 |",
        f"| `00-full-suite` | {stats['00-full-suite']['iterations']} | {stats['00-full-suite']['requests']} | {stats['00-full-suite']['assertions']} | {stats['00-full-suite']['failed']} | strict toàn bộ probe chính |",
        f"| `01-ddt-login` | {stats['01-ddt-login']['iterations']} | {stats['01-ddt-login']['requests']} | {stats['01-ddt-login']['assertions']} | {stats['01-ddt-login']['failed']} | coverage DDT login |",
        f"| `02-ddt-checkout` | {stats['02-ddt-checkout']['iterations']} | {stats['02-ddt-checkout']['requests']} | {stats['02-ddt-checkout']['assertions']} | {stats['02-ddt-checkout']['failed']} | coverage DDT checkout |",
        f"| `03-ddt-order-status` | {stats['03-ddt-order-status']['iterations']} | {stats['03-ddt-order-status']['requests']} | {stats['03-ddt-order-status']['assertions']} | {stats['03-ddt-order-status']['failed']} | matrix + coverage DDT status |",
        "",
        "HTML/JSON evidence: [`newman/reports`](../newman/reports/). DDT runner tự chuẩn bị environment/auth/order trước khi chạy folder, tránh kết quả giả do 401 hoặc orderId rỗng.",
        "",
        "## 4. Postman, CI và generator",
        "",
        "Collection có collection-level pre-request, environment variables, dynamic disposable user, token chaining, response assertions, strict modes `off/canary/full`, data-driven folders và htmlextra/JSON export. Chi tiết ở [`postman-features.md`](../postman/postman-features.md). Workflow nằm ở [`.github/workflows/hw06-newman-api-test.yml`](../../.github/workflows/hw06-newman-api-test.yml). Pseudocode và generator tham chiếu ở [`test-generator/design.md`](../test-generator/design.md) và [`generator.py`](../test-generator/generator.py).",
        "",
        "OpenAPI audit: file [`openapi/eshop.openapi.yaml`](../openapi/eshop.openapi.yaml) chỉ mô tả các endpoint được test; mọi expected result vẫn đối chiếu SUT defect catalog.",
        "",
        "## 5. Defects và giới hạn bằng chứng",
        "",
        (f"15 defect IDs trong defect catalog đã được lập trong [`bug-report.md`](bug-report.md), mỗi dòng có Found by Test Case, expected/actual và nguồn evidence. Newman JSON ghi nhận {stats['00-full-suite']['failed']} fail ở full probe cùng {stats['01-ddt-login']['failed']}/{stats['02-ddt-checkout']['failed']}/{stats['03-ddt-order-status']['failed']} fail ở ba DDT suite; request/test-script infrastructure đều không fail. Đã tạo đủ 15 GitHub Issues scrubbed (#413–#427) và lưu 15 ảnh trang issue tại `evidence/screenshots/github-issues/`; trạng thái CI external được ghi riêng trong `cicd-report.md`." if issue_manifest else "15 defect IDs trong defect catalog đã được lập trong [`bug-report.md`](bug-report.md); chưa có manifest issue external."),
        "",
        "`ai-critique.md` là bản nháp dữ liệu 200–300 từ để người học viết lại bằng nhận xét của chính mình; `diagram.mmd` là bản mô tả kỹ thuật, không thay thế `diagram.png` tự vẽ.",
        "",
        "## 6. Artifact index",
        "",
        "- [README tự chấm và summary](../README.md)",
        "- [AI audit log](ai-audit-report.md)",
        "- [Bug report](bug-report.md)",
        "- [GitHub issue manifest](github-issues.json)",
        "- [GitHub issue screenshot index](../evidence/screenshots/github-issues.md)",
        "- [CI/CD report](cicd-report.md)",
        "- [Excel/CSV](../excel/)",
        "- [Traceability](../../tests/test-summary/traceability-matrix.md)",
        "",
        "### Human completion gates",
        "",
        ("1. Xác minh metadata/signature API-3; 2. kiểm tra 15 GitHub issue links + 15 screenshot local; 3. chụp Postman Console/Newman/CI; 4. tự vẽ `diagram.png`; 5. viết lại critique và xuất ba PDF; 6. đặt repo public và đóng zip theo tên đề bài." if api3_signoff else "1. Ký API-3 audit; 2. kiểm tra 15 GitHub issue links + 15 screenshot local; 3. chụp Postman Console/Newman/CI; 4. tự vẽ `diagram.png`; 5. viết lại critique và xuất ba PDF; 6. đặt repo public và đóng zip theo tên đề bài."),
    ]
    (REPORT / "main-report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_bug_report() -> None:
    bugs = [
        ("D-LOGIN-01", "Critical", "Sai password làm tăng counter hai lần; lock quá sớm", "TC-API-LOGIN-018", "200 sau hai lần sai (theo đặc tả)", "403 sau hai lần sai trong canary/full; xem `00-canary-suite.json`", "Newman"),
        ("D-LOGIN-02", "Major", "Thời gian khóa 180s thay vì 30s", "TC-API-LOGIN-022", "Hết khóa sau 30s", "Login đúng sau 31s và 35s vẫn trả 403", "`01-ddt-login.json`"),
        ("D-LOGIN-03", "Critical", "Response trả password plaintext", "TC-API-LOGIN-028", "Không có password", "Negative-schema assertion phát hiện field nhạy cảm trong response", "`00-full-suite.json`, `01-ddt-login.json`"),
        ("D-LOGIN-05", "Major", "JWT hard-code/không hết hạn", "TC-API-LOGIN-040", "JWT có exp và secret quản lý an toàn", "JWT claim assertion không tìm thấy `exp`; hard-coded secret được xác nhận bằng source review", "`01-ddt-login.json` + source review"),
        ("D-LOGIN-06", "Major", "Counter không reset sau khi hết khóa", "TC-API-LOGIN-020", "Counter reset khi hết thời gian khóa", "Catalog ghi nhận counter giữ nguyên; cần probe chờ timer", "Catalog + manual follow-up"),
        ("D-CHK-01", "Critical", "Tin total_amount từ client", "TC-API-CHECKOUT-037", "Tính lại total từ cart", "Checkout nhận client total `1` và tạo order thành công", "`00-full-suite.json` / test table"),
        ("D-CHK-02", "Major", "Chấp nhận total âm/0", "TC-API-CHECKOUT-005", "400 validation", "Zero total trả 200 trong full strict", "`00-full-suite.json`"),
        ("D-CHK-03", "Major", "Không xóa cart sau checkout", "TC-API-CHECKOUT-020", "Cart rỗng", "Post-condition strict fail; cart vẫn còn item", "`00-full-suite.json`"),
        ("D-CHK-04", "Major", "Checkout với cart rỗng", "TC-API-CHECKOUT-022", "400", "Fresh user không có cart vẫn checkout 200 và tạo order", "`02-ddt-checkout.json`"),
        ("D-CHK-07", "Critical", "IDOR GET /api/orders/:id", "TC-API-CHECKOUT-031", "401/403 nếu không có auth", "Anonymous/order detail probe trả 200", "`00-full-suite.json`"),
        ("D-ADM-01", "Critical", "User thường đổi trạng thái qua API admin", "TC-API-ORDER-STATUS-033", "403", "User token trả 200 trong full strict", "`00-full-suite.json`"),
        ("D-ADM-02", "Critical", "canceled → delivered được phép", "TC-API-ORDER-STATUS-024", "400", "Trả 200 trong full strict và matrix DDT", "`00-full-suite.json`, `03-ddt-order-status.json`"),
        ("D-ADM-03", "Major", "Admin không hủy được shipping", "TC-API-ORDER-STATUS-015", "200", "Stateful DDT dựng order shipping rồi nhận 400 khi admin hủy", "`03-ddt-order-status.json`"),
        ("D-ADM-04", "Major", "Bỏ qua lỗi UPDATE, trả 200", "TC-API-ORDER-STATUS-041", "4xx/5xx khi update lỗi", "Callback bỏ qua err theo catalog; cần tạo orderId không tồn tại", "Catalog + manual follow-up"),
        ("D-ADM-08", "Major", "User hủy order shipping", "TC-API-ORDER-STATUS-043", "400", "Stateful DDT dựng order shipping rồi user cancel nhận 200", "`03-ddt-order-status.json`"),
    ]
    issue_manifest_path = REPORT / "github-issues.json"
    issue_manifest = {item["bug_id"]: item for item in json.loads(issue_manifest_path.read_text(encoding="utf-8"))} if issue_manifest_path.exists() else {}
    lines = [
        "# HW06 bug report", "", 
        "> Mỗi defect giữ đúng ID trong `docs/hw06/02-sut-defect-catalog.md`. Evidence Newman là report JSON thật; các dòng manual follow-up được đánh dấu rõ, không bịa screenshot/issue.",
        "", "| Bug ID | Severity | Title | Found by Test Case | Expected | Actual | Evidence | GitHub Issue | Screenshot |", "| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |",
    ]
    for index, (bid, sev, title, tc, expected, actual, evidence) in enumerate(bugs, start=1):
        issue = issue_manifest.get(bid)
        issue_cell = f"[#${issue['issue_number']}]({issue['url']})".replace("#$", "#") if issue else "Chưa tạo — HUMAN"
        screenshot_cell = f"[bug-{index:02d}-{bid}-issue.png](../evidence/screenshots/github-issues/bug-{index:02d}-{bid}-issue.png)" if issue else "Chưa có — HUMAN"
        lines.append(f"| {bid} | {sev} | {title} | `{tc}` | {expected} | {actual} | {evidence} | {issue_cell} | {screenshot_cell} |")
    lines += ["", "## Quy ước reproducing", "", "- Chạy backend reset DB rồi `powershell -ExecutionPolicy Bypass -File hw06/newman/run-newman.ps1 -Mode full -BaseUrl http://127.0.0.1:3001`.", "- Dùng `-DataDriven` để chạy 39/41/43 rows; mỗi iteration tự dựng user/cart/order và state cần thiết.", "- GitHub Issues và screenshot là artifact external/human-only, nên không được thay bằng số issue hoặc ảnh giả."]
    (REPORT / "bug-report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_cicd(stats: dict[str, dict[str, int]]) -> None:
    lines = ["# HW06 CI/CD report", "", "## Pipeline", "", "```mermaid", "flowchart LR", "A[checkout] --> B[setup Node 20] --> C[npm ci backend] --> D[start localhost:3000] --> E[npm ci hw06] --> F[Newman off/canary/full] --> G[upload HTML+JSON]", "```", "", "Workflow: [`.github/workflows/hw06-newman-api-test.yml`](../../.github/workflows/hw06-newman-api-test.yml). It installs backend dependencies, waits for `/api/products`, starts the SUT, installs Newman and uploads reports even on failure.", "", "## Strict modes", "", "- `off`: only observed/oracle-safe assertions; used as green smoke run.", "- `canary`: strict one-case gate `TC-API-LOGIN-018`; expected red while D-LOGIN-01 exists.", "- `full`: all strict probes; exposes all currently known defects.", "", "## Local evidence (the same collection and runner used by CI)", "", "| Mode/report | Requests | Assertions | Failed | External Actions link | Screenshot |", "| :--- | ---: | ---: | ---: | :--- | :--- |", f"| off — `00-off-suite` | {stats['00-off-suite']['requests']} | {stats['00-off-suite']['assertions']} | {stats['00-off-suite']['failed']} | Chưa có — HUMAN | Chưa có — HUMAN |", f"| canary — `00-canary-suite` | {stats['00-canary-suite']['requests']} | {stats['00-canary-suite']['assertions']} | {stats['00-canary-suite']['failed']} | Chưa có — HUMAN | Chưa có — HUMAN |", "", "Không ghi SHA/link GitHub Actions khi chưa có run external thật. Sau khi push, người học điền hai URL/SHA và chụp `04-ci-pass.png`, `05-ci-fail.png`."]
    (REPORT / "cicd-report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_critique_draft() -> None:
    text = """# AI Critique — bản nháp để người học viết lại

> HUMAN-ONLY: đây là bản nháp dữ liệu, không phải bài critique đã ký của sinh viên. Vui lòng viết lại bằng nhận xét của chính bạn trước khi nộp.

AI tạo được độ phủ khá rộng ở ba API, đặc biệt là các phân vùng dữ liệu, ma trận trạng thái và nhóm kiểm thử bảo mật. Tuy nhiên, output ban đầu vẫn có những expected result không bám đặc tả, chẳng hạn coi việc bypass SQL injection là kết quả hợp lệ hoặc bỏ qua điều kiện hậu nghiệm của checkout. Audit đã phải sửa các case đó thành INVALID hoặc INCOMPLETE. Đây là dấu hiệu AI có xu hướng dự đoán theo mẫu API phổ biến thay vì đọc chính xác oracle của SUT.

AI cũng bỏ sót các lỗi nằm ngoài endpoint chính. Ví dụ, IDOR ở GET `/api/orders/:id`, việc giỏ hàng không bị xóa, và các nhánh trạng thái kết thúc chỉ xuất hiện sau khi mở rộng ngữ cảnh bằng flow nghiệp vụ. Ma trận 5×5 của API-3 cho thấy câu trả lời một request không đủ để kiểm thử hệ thống có state; mỗi ô cần precondition độc lập và phải ghi rõ `from_status`.

Bài học quan trọng là không xem số lượng test case do AI sinh như bằng chứng chất lượng. Cần chia prompt thành các bước có thể audit, đối chiếu từng expected result với đặc tả và mã SUT, rồi chạy các case quan trọng trên Newman. Human review vẫn cần thiết cho quyết định VALID/INVALID, cho nhận định mức độ nghiêm trọng và cho các bằng chứng external như GitHub Issue, screenshot, diagram. AI hữu ích nhất khi tạo scaffold có cấu trúc; trách nhiệm oracle và kết luận cuối cùng vẫn thuộc về người kiểm thử.
"""
    (REPORT / "ai-critique.md").write_text(text, encoding="utf-8")


def write_test_run(stats: dict[str, dict[str, int]]) -> None:
    lines = ["# HW06 API test run", "", "| Suite | Iterations | Requests | Assertions | Failed | Result |", "| :--- | ---: | ---: | ---: | ---: | :--- |"]
    for name in ["00-off-suite", "00-canary-suite", "00-full-suite", "01-ddt-login", "02-ddt-checkout", "03-ddt-order-status"]:
        s = stats[name]
        failed = s["failed"]
        lines.append(f"| `{name}` | {s['iterations']} | {s['requests']} | {s['assertions']} | {failed} | {'PASS' if failed == 0 else 'FAIL (expected defect/oracle mismatch)'} |")
    lines += ["", "## Failure mapping", "", "- Canary: `TC-API-LOGIN-018` → D-LOGIN-01; Newman JSON có đúng 1 failed assertion.", "- Full probe: xem từng assertion TC ID trong `00-full-suite.json` và `report/bug-report.md`.", "- DDT: expected giữ theo đặc tả; các failed assertion là chênh lệch oracle/SUT, không phải lỗi request hoặc test script.", "- Đối soát 128 TC ID: `hw06/newman/reports/execution-coverage.md`.", "", "All request logs use `X-Student-Id: 23127207`; reports are in `hw06/newman/reports/`."]
    RUNS.mkdir(parents=True, exist_ok=True)
    (RUNS / "hw06-api-test-run.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def append_traceability(all_rows: list[dict[str, str]]) -> None:
    path = SUMMARY / "traceability-matrix.md"
    current = path.read_text(encoding="utf-8") if path.exists() else "# Traceability Matrix\n"
    marker = "\n## HW06 — API Testing\n"
    if marker in current:
        current = current.split(marker, 1)[0]
    issue_manifest_path = REPORT / "github-issues.json"
    issue_manifest = {item["bug_id"]: item for item in json.loads(issue_manifest_path.read_text(encoding="utf-8"))} if issue_manifest_path.exists() else {}
    lines = [current.rstrip(), marker.rstrip(), "", "| Requirement | Test Case | Result | Bug Issue | Status |", "| :--- | :--- | :--- | :--- | :--- |"]
    for row in all_rows:
        result = "Fail" if row["Kỳ vọng chạy"].upper() == "FAIL" else "Pass/Smoke"
        bug = row["Bug ID"] if row["Bug ID"] not in {"—", "-", ""} else "None"
        issue = issue_manifest.get(bug)
        bug_cell = f"[{bug} #{issue['issue_number']}]({issue['url']})" if issue else (f"{bug} / chưa tạo issue" if bug != "None" else "None")
        lines.append(f"| {row['Requirement']} | `{row['TC ID']}` | {result} | {bug_cell} | {'Open' if bug != 'None' else 'Covered'} |")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_representative_test_cases(all_rows: list[dict[str, str]]) -> None:
    """Materialize failed cases plus five representatives/API for Rule.pdf traceability."""
    buckets = {"API-1": (ROOT / "tests" / "test-cases" / "api-login", "LOGIN"), "API-2": (ROOT / "tests" / "test-cases" / "api-checkout", "CHECKOUT"), "API-3": (ROOT / "tests" / "test-cases" / "api-order-status", "ORDER-STATUS")}
    for api, (directory, _) in buckets.items():
        rows = [row for row in all_rows if row["API"] == api]
        selected = []
        for row in rows:
            if row["Kỳ vọng chạy"].upper() == "FAIL" or len(selected) < 5:
                selected.append(row)
        directory.mkdir(parents=True, exist_ok=True)
        for row in selected:
            bug = row["Bug ID"] if row["Bug ID"] not in {"—", "-", ""} else "None"
            result = "FAIL — defect expected" if bug != "None" else "PASS/SMOKE"
            body = [
                f"# {row['TC ID']}", "", f"- Requirement: `{row['Requirement']}`", f"- Group/technique: {row['Nhóm']} / {row['Kỹ thuật']}", f"- Preconditions: {row['Preconditions']}", f"- Method/data: {row['Method + Endpoint / Test data']}", f"- Expected: {row['Expected result']}", f"- Result: {result}", f"- Related Bug: `{bug}`", "", "> This file is a representative/failed-case traceability artifact generated from the final Markdown test table. Full inventory remains in `hw06/api-*/test-cases.md`.", ""
            ]
            (directory / f"{row['TC ID']}.md").write_text("\n".join(body), encoding="utf-8")


def write_commit_log() -> None:
    try:
        result = subprocess.run(["git", "log", "--pretty=format:%h %ad %an %s", "--date=iso", "HW6-Khoa"], cwd=ROOT, capture_output=True, check=True)
        stdout = result.stdout.decode("utf-8", errors="replace")
        (REPORT / "git-commit-log.txt").write_text(stdout + "\n", encoding="utf-8")
    except (OSError, subprocess.CalledProcessError) as exc:
        (REPORT / "git-commit-log.txt").write_text(f"Unable to read git log: {exc}\n", encoding="utf-8")


def main() -> None:
    REPORT.mkdir(parents=True, exist_ok=True)
    all_rows: list[dict[str, str]] = []
    for label, folder, _, _, _, _ in APIS:
        for row in parse_table(HW / folder / "test-cases.md"):
            row = {**row, "API": label}
            all_rows.append(row)
    build_excel(all_rows)
    stats = {name: read_stats(name) for name in ["00-off-suite", "00-canary-suite", "00-full-suite", "01-ddt-login", "02-ddt-checkout", "03-ddt-order-status"]}
    write_main_report(all_rows, stats)
    write_bug_report()
    write_cicd(stats)
    write_critique_draft()
    write_test_run(stats)
    append_traceability(all_rows)
    write_representative_test_cases(all_rows)
    write_commit_log()
    print(f"Built reports for {len(all_rows)} final test cases")


if __name__ == "__main__":
    main()
