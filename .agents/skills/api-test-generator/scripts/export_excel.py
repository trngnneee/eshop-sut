#!/usr/bin/env python3
"""
export_excel.py — test_cases.json -> test_cases.xlsx

Writes one row per test case plus a Summary sheet tallying counts by stage
and by audit label (matches the assignment's "Excel test cases and test
summary" submission item).

Requires: openpyxl (pip install openpyxl --break-system-packages)

Usage:
    python3 export_excel.py test_cases.json --out test_cases.xlsx
"""
import argparse
import json
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("id", "ID"),
    ("stage", "Stage"),
    ("endpoint", "Endpoint"),
    ("title", "Title"),
    ("category", "Category"),
    ("priority", "Priority"),
    ("fr_ref", "FR Ref"),
    ("sec_ref", "SEC Ref"),
    ("preconditions", "Preconditions"),
    ("request", "Request"),
    ("expected", "Expected"),
    ("audit_label", "Audit Label"),
    ("audit_reasoning", "Audit Reasoning"),
    ("why_ai_missed", "Why AI Missed (extensions only)"),
]


def flatten(tc):
    audit = tc.get("audit") or {}
    return {
        "id": tc.get("id"),
        "stage": tc.get("stage"),
        "endpoint": tc.get("endpoint"),
        "title": tc.get("title"),
        "category": tc.get("category"),
        "priority": tc.get("priority"),
        "fr_ref": tc.get("fr_ref"),
        "sec_ref": tc.get("sec_ref"),
        "preconditions": tc.get("preconditions"),
        "request": json.dumps(tc.get("request", {}), ensure_ascii=False),
        "expected": json.dumps(tc.get("expected", {}), ensure_ascii=False),
        "audit_label": audit.get("label"),
        "audit_reasoning": audit.get("reasoning"),
        "why_ai_missed": tc.get("why_ai_missed"),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("test_cases_json")
    ap.add_argument("--out", default="test_cases.xlsx")
    args = ap.parse_args()

    with open(args.test_cases_json, "r", encoding="utf-8") as f:
        cases = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, tc in enumerate(cases, start=2):
        flat = flatten(tc)
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=flat.get(key))

    for col_idx, (key, label) in enumerate(COLUMNS, start=1):
        width = 40 if key in ("request", "expected", "title", "preconditions", "audit_reasoning", "why_ai_missed") else 16
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    stage_counts = Counter(tc.get("stage") for tc in cases)
    label_counts = Counter((tc.get("audit") or {}).get("label") for tc in cases)

    ws2.append(["Total test cases", len(cases)])
    ws2.append([])
    ws2.append(["By stage"])
    for stage, count in sorted(stage_counts.items(), key=lambda x: (-x[1], str(x[0]))):
        ws2.append([stage or "(unset)", count])
    ws2.append([])
    ws2.append(["By audit label"])
    for label, count in sorted(label_counts.items(), key=lambda x: (-x[1], str(x[0]))):
        ws2.append([label or "(unaudited)", count])
    for col in ("A", "B"):
        ws2.column_dimensions[col].width = 24

    wb.save(args.out)
    print(f"Wrote {args.out} ({len(cases)} test cases)")


if __name__ == "__main__":
    main()
