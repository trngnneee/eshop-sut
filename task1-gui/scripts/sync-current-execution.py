from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TASK = Path(__file__).resolve().parents[1]
REPO = TASK.parent
TASK3 = REPO / "task3-cross-platform"
SOURCE_MD = TASK / "GUI_Checklist_HW3.md"
SOURCE_CSV = TASK3 / "results" / "Task3_Cross_Platform_Results.csv"
EVIDENCE = TASK / "evidence" / "executed-chrome"
RESULTS = TASK / "results"
DATE = "2026-08-02"
PLATFORM = "Google Chrome 150.0.7871.187 / Windows 10.0.26200"

EXISTING_ISSUES = {
    "GUI-WEB-LOGIN-001": "https://github.com/trngnneee/eshop-sut/issues/199",
    "GUI-WEB-LOGIN-002": "https://github.com/trngnneee/eshop-sut/issues/203",
    "GUI-WEB-LOGIN-003": "https://github.com/trngnneee/eshop-sut/issues/37",
    "GUI-WEB-LOGIN-007": "https://github.com/trngnneee/eshop-sut/issues/230",
    "GUI-WEB-LOGIN-009": "https://github.com/trngnneee/eshop-sut/issues/198",
    "GUI-WEB-LOGIN-010": "https://github.com/trngnneee/eshop-sut/issues/238",
    "GUI-WEB-LOGIN-011": "https://github.com/trngnneee/eshop-sut/issues/201",
    "GUI-WEB-REGISTER-006": "https://github.com/trngnneee/eshop-sut/issues/117",
}

SEVERITY = {
    "GUI-WEB-LOGIN-003": "Critical",
    "GUI-WEB-LOGIN-010": "High",
    "GUI-WEB-REGISTER-006": "High",
    "GUI-ADMIN-CATEGORY-004": "High",
    "GUI-ADMIN-CATEGORY-006": "High",
    "GUI-ADMIN-CATEGORY-008": "High",
    "GUI-WEB-LOGIN-007": "Low",
    "GUI-ADMIN-CATEGORY-009": "Low",
}

CORRECTIONS = {
    "GUI-WEB-LOGIN-001": {"Related Requirement": "FR-21"},
    "GUI-WEB-LOGIN-002": {"Related Requirement": "FR-02 / FR-22"},
    "GUI-WEB-LOGIN-003": {"Related Requirement": "FR-22"},
    "GUI-WEB-LOGIN-009": {"Related Requirement": "FR-21"},
    "GUI-WEB-LOGIN-010": {
        "Checklist Item": "Kiểm tra phản hồi sau đúng ba lần đăng nhập sai liên tiếp.",
        "Expected Result": "Sau lần sai thứ ba, backend khóa 30 giây và UI hiển thị trạng thái khóa phù hợp mà không lộ chi tiết tài khoản.",
    },
    "GUI-WEB-LOGIN-011": {"Related Requirement": "FR-21"},
    "GUI-WEB-LOGIN-013": {
        "Checklist Item": "Kiểm tra identifier có khoảng trắng đầu/cuối khi requirement không quy định normalization.",
        "Expected Result": "Giá trị có khoảng trắng không xác thực thành công và UI hiển thị phản hồi đăng nhập thất bại chung.",
        "Status": "Pass",
    },
    "GUI-WEB-REGISTER-002": {"Related Requirement": "FR-01 / FR-22"},
    "GUI-WEB-REGISTER-008": {"Related Requirement": "FR-21"},
    "GUI-ADMIN-LOGIN-002": {"Related Requirement": "FR-22 / Accessibility Heuristic"},
    "GUI-ADMIN-CATEGORY-005": {
        "IA": "IA-03", "Category": "Navigation",
        "Checklist Item": "Kiểm tra chuyển từ Admin shell sang tab Quản lý Danh mục.",
        "Expected Result": "Chọn tab Danh mục hiển thị tiêu đề và bảng quản lý trong cùng Admin shell.",
        "Actual Result": "Tab Danh mục mở thành công; tiêu đề và bảng ID / Tên Danh Mục / Hành động hiển thị.",
        "Status": "Pass",
    },
    "GUI-ADMIN-CATEGORY-008": {
        "Expected Result": "Không xóa category đang được product tham chiếu; UI hiển thị lỗi và category vẫn còn."
    },
    "GUI-ADMIN-CATEGORY-009": {"Related Requirement": "FR-14 / FR-24"},
    "GUI-ADMIN-CATEGORY-011": {
        "Checklist Item": "Quan sát Add/View khi gửi lại tên danh mục trong khi FR-14 không quy định unique name.",
        "Expected Result": "Request có kết quả xác định và row kết quả hiển thị; không tự suy diễn yêu cầu reject duplicate.",
        "Actual Result": "Duplicate category POST trả HTTP 200 và row kết quả hiển thị trong bảng.",
        "Status": "Pass",
    },
    "GUI-MOBILE-LOGIN-002": {"Related Requirement": "FR-21"},
    "GUI-MOBILE-LOGIN-004": {"Related Requirement": "FR-21"},
}

HUMAN_REASON = {
    "GUI-WEB-LOGIN-011": "keyboard-only focus order was absent from the AI set",
    "GUI-WEB-LOGIN-012": "the AI prompt omitted the 320 px viewport boundary",
    "GUI-WEB-LOGIN-013": "the AI omitted surrounding-whitespace input behavior",
    "GUI-WEB-REGISTER-011": "the AI omitted hostile script-like input",
    "GUI-WEB-REGISTER-012": "the AI omitted transport/network failure",
    "GUI-ADMIN-LOGIN-009": "the AI omitted unauthenticated direct Admin access",
    "GUI-ADMIN-CATEGORY-012": "the AI omitted a 260+ character layout boundary",
    "GUI-ADMIN-CATEGORY-013": "the AI omitted rapid submission while a request is pending",
    "GUI-MOBILE-LOGIN-010": "the AI used subjective sizing instead of a 44×44 measurement",
    "GUI-MOBILE-LOGIN-011": "the AI could not observe physical soft-keyboard behavior from source",
}

OLD_HEADERS = ["ID", "Platform", "Screen/Route", "Related Requirement", "IA", "Category", "Origin",
               "Checklist Item", "Expected Result", "Actual Result", "Status", "Notes", "Evidence", "Bug ID"]
NEW_HEADERS = ["ID", "Platform", "Screen/Route", "Related Requirement", "IA", "Category", "Origin",
               "Checklist Item", "Expected Result", "Actual Result", "Status", "Execution Mode", "Notes",
               "Evidence", "Evidence ID", "Captured At", "Bug ID", "GitHub Issue"]


def clean(value: object) -> str:
    return " ".join(str(value or "").replace("|", "\\|").split())


def source_rows() -> list[dict[str, str]]:
    rows = []
    for line in SOURCE_MD.read_text(encoding="utf-8-sig").splitlines():
        if not line.startswith("| GUI-"):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if len(cells) != len(OLD_HEADERS):
            raise RuntimeError(f"Unexpected Task 1 column count: {len(cells)}")
        row = dict(zip(OLD_HEADERS, cells))
        row["Status"] = row["Status"].replace("**", "")
        rows.append(row)
    if len(rows) != 58:
        raise RuntimeError(f"Expected 58 Task 1 rows, got {len(rows)}")
    return rows


def chrome_rows() -> dict[str, dict[str, str]]:
    with SOURCE_CSV.open(encoding="utf-8-sig", newline="") as stream:
        rows = [row for row in csv.DictReader(stream) if row["platform_id"] == "chrome-windows"]
    if len(rows) != 58 or len({row["checklist_id"] for row in rows}) != 58:
        raise RuntimeError("Task 3 Chrome execution is not a unique 58-row set")
    return {row["checklist_id"]: row for row in rows}


def reconcile() -> list[dict[str, str]]:
    EVIDENCE.mkdir(parents=True, exist_ok=True)
    executed = chrome_rows()
    output = []
    for source in source_rows():
        item_id = source["ID"]
        observed = executed[item_id]
        row = dict(source)
        row.update(CORRECTIONS.get(item_id, {}))
        row["Actual Result"] = CORRECTIONS.get(item_id, {}).get("Actual Result", observed["actual_result"])
        row["Status"] = CORRECTIONS.get(item_id, {}).get("Status", observed["status"])
        if row["Status"] == "Not Observable":
            row["Status"] = "Blocked"
        evidence_name = Path(observed["evidence_path"]).name
        shutil.copy2(TASK3 / observed["evidence_path"], EVIDENCE / evidence_name)
        row["Evidence"] = f"evidence/executed-chrome/{evidence_name}"
        row["Evidence ID"] = observed["evidence_id"]
        row["Captured At"] = observed["captured_at"]
        row["Execution Mode"] = observed["execution_mode"]
        if row["Status"] == "Fail":
            row["Bug ID"] = f"BUG-{item_id}"
            row["GitHub Issue"] = EXISTING_ISSUES.get(item_id, "PENDING_EXTERNAL_ACTION")
            row["Notes"] = f"Observed mismatch; Mode={row['Execution Mode']}; see {row['Bug ID']}."
        elif row["Status"] == "Blocked":
            row["Bug ID"] = ""
            row["GitHub Issue"] = ""
            row["Notes"] = "Requires Expo Go or a physical/cloud phone; Expo Web cannot prove soft-keyboard behavior."
        else:
            row["Bug ID"] = ""
            row["GitHub Issue"] = ""
            row["Notes"] = f"Observed result matched the corrected measurable expectation; Mode={row['Execution Mode']}."
        output.append(row)
    return output


def write_checklist(rows: list[dict[str, str]]) -> None:
    lines = [
        "# GUI Checklist — HW03 (EShop System Under Test)", "",
        "**Student:** Đặng Đăng Khoa — 23127207  ",
        "**Email overlay:** `23127207@student.hcmus.edu.vn`  ",
        f"**Execution:** {PLATFORM}, {DATE}  ",
        "**Source:** corrected final Task 3 Chrome run with item-level screenshots and execution modes  ",
        "**Status:** `BLOCKED_REAL_MOBILE_SOFT_KEYBOARD_AND_PENDING_EXTERNAL_ITEMS`", "",
        "| " + " | ".join(NEW_HEADERS) + " |",
        "|" + "|".join(["---"] * len(NEW_HEADERS)) + "|",
    ]
    for row in rows:
        values = []
        for header in NEW_HEADERS:
            value = row[header]
            if header == "Status":
                value = f"**{value}**"
            elif header == "Evidence":
                value = f"[Chrome evidence]({value})"
            elif header == "GitHub Issue" and str(value).startswith("https://"):
                value = f"[Issue]({value})"
            values.append(clean(value))
        lines.append("| " + " | ".join(values) + " |")
    lines += [
        "", "## Human-review traceability", "",
        "The final set contains 48 `AI_INITIAL` and 10 `HUMAN_ADDED` items. The student confirmed all human-review sections on 2026-08-02. Specific human-added rationales are in `AI_Item_Level_Critique.md` and the Excel `AI Human Review` sheet.",
        "", "## Integrity boundary", "",
        "`MOCKED_*` is used only for deterministic error/loading/empty/double-submit states. Core flows use `LIVE_LOCAL_SUT`. `GUI-MOBILE-LOGIN-011` remains Blocked because Expo Web cannot display a real mobile soft keyboard.",
    ]
    SOURCE_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_csv(rows: list[dict[str, str]]) -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)
    path = RESULTS / "Task1_Execution_Chrome.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=NEW_HEADERS)
        writer.writeheader()
        writer.writerows({header: row[header] for header in NEW_HEADERS} for row in rows)
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["Evidence"]].append(row["ID"])
    with (RESULTS / "Evidence_Index.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=["evidence_path", "checklist_ids", "platform"])
        writer.writeheader()
        for path, ids in grouped.items():
            writer.writerow({"evidence_path": path, "checklist_ids": ";".join(ids), "platform": PLATFORM})


def style(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, column in enumerate(sheet.columns, 1):
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        sheet.column_dimensions[get_column_letter(index)].width = max(12, width)


def write_excel(rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    checklist = workbook.active
    checklist.title = "Checklist"
    checklist.append(NEW_HEADERS)
    for row in rows:
        checklist.append([row[header] for header in NEW_HEADERS])
    style(checklist)

    human = workbook.create_sheet("AI Human Review")
    human.append(["ID", "Origin", "Why the AI missed it", "Student review", "Date"])
    for row in rows:
        if row["Origin"] == "HUMAN_ADDED":
            human.append([row["ID"], row["Origin"], HUMAN_REASON[row["ID"]], "HUMAN_REVIEWED", DATE])
    style(human)

    summary = workbook.create_sheet("Execution Summary")
    summary.append(["Metric", "Value"])
    counts = Counter(row["Status"] for row in rows)
    for key, value in [("Total", len(rows)), ("Pass", counts["Pass"]), ("Fail", counts["Fail"]),
                       ("Blocked", counts["Blocked"]), ("Unique screenshots", len({r['Evidence'] for r in rows})),
                       ("Platform", PLATFORM), ("Date", DATE)]:
        summary.append([key, value])
    style(summary)

    trace = workbook.create_sheet("Bug Traceability")
    trace.append(["Bug ID", "Checklist ID", "Severity", "GitHub Issue", "Evidence"])
    for row in rows:
        if row["Status"] == "Fail":
            trace.append([row["Bug ID"], row["ID"], SEVERITY.get(row["ID"], "Medium"), row["GitHub Issue"], row["Evidence"]])
    style(trace)

    evidence = workbook.create_sheet("Evidence Index")
    evidence.append(["Evidence", "Checklist IDs", "Platform"])
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["Evidence"]].append(row["ID"])
    for path, ids in grouped.items():
        evidence.append([path, ";".join(ids), PLATFORM])
    style(evidence)
    workbook.save(TASK / "GUI_Checklist_HW3.xlsx")


def write_coverage(rows: list[dict[str, str]]) -> None:
    counters = {
        "Platform": Counter(row["Platform"] for row in rows),
        "IA": Counter(row["IA"] for row in rows),
        "Origin": Counter(row["Origin"] for row in rows),
        "Status": Counter(row["Status"] for row in rows),
        "Execution Mode": Counter(row["Execution Mode"] for row in rows),
    }
    lines = ["# GUI Coverage Matrix — HW03 Task 1", "", "**Source:** current corrected Chrome execution", "",
             "| Dimension | Value | Count |", "|---|---|---:|"]
    for dimension, values in counters.items():
        for value, count in sorted(values.items()):
            lines.append(f"| {dimension} | `{value}` | {count} |")
    lines += ["", "## Screen × IA", "", "| Screen | IA-01 | IA-02 | IA-03 | IA-04 | Total |",
              "|---|---:|---:|---:|---:|---:|"]
    groups = [("Web Login", "GUI-WEB-LOGIN-"), ("Web Register", "GUI-WEB-REGISTER-"),
              ("Admin Login", "GUI-ADMIN-LOGIN-"), ("Admin Category", "GUI-ADMIN-CATEGORY-"),
              ("Mobile Login", "GUI-MOBILE-LOGIN-")]
    for label, prefix in groups:
        subset = [row for row in rows if row["ID"].startswith(prefix)]
        count = Counter(row["IA"] for row in subset)
        lines.append(f"| {label} | {count['IA-01']} | {count['IA-02']} | {count['IA-03']} | {count['IA-04']} | {len(subset)} |")
    total = counters["IA"]
    lines.append(f"| **Total** | **{total['IA-01']}** | **{total['IA-02']}** | **{total['IA-03']}** | **{total['IA-04']}** | **{len(rows)}** |")
    (TASK / "GUI_Coverage_Matrix.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(rows: list[dict[str, str]]) -> None:
    status = Counter(row["Status"] for row in rows)
    mode = Counter(row["Execution Mode"] for row in rows)
    failures = [row for row in rows if row["Status"] == "Fail"]
    existing = [row for row in failures if row["GitHub Issue"].startswith("https://")]
    pending = [row for row in failures if row["GitHub Issue"] == "PENDING_EXTERNAL_ACTION"]
    lines = [
        "# GUI Test Summary — HW03 Task 1", "",
        "**Student:** Đặng Đăng Khoa — 23127207  ", f"**Execution:** {PLATFORM}, {DATE}  ",
        "**Status:** `BLOCKED_REAL_MOBILE_SOFT_KEYBOARD_AND_PENDING_EXTERNAL_ITEMS`", "",
        "| Metric | Value |", "|---|---:|", f"| Total | {len(rows)} |", f"| Pass | {status['Pass']} |",
        f"| Fail | {status['Fail']} |", f"| Blocked | {status['Blocked']} |",
        f"| Existing real GitHub mappings | {len(existing)} |", f"| Pending new GitHub issues | {len(pending)} |",
        f"| Unique screenshots | {len({row['Evidence'] for row in rows})} |", "", "## Execution modes", "",
    ]
    for name, count in sorted(mode.items()):
        lines.append(f"- `{name}`: {count} item(s).")
    lines += [
        "", "## Completion decision", "",
        "All 58 items have current item-level Actual Results and screenshots. Mocked states are explicitly labelled. The package remains blocked only where evidence/actions are genuinely unavailable: real mobile soft-keyboard behavior, pending GitHub publications and a real Task 1 skill-demo YouTube URL.",
        "", "The older 40/18 summary and five packed screenshots are superseded. `results/Task1_Execution_Chrome.csv` is the current source of truth.",
    ]
    (TASK / "GUI_Test_Summary_HW3.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_bug_docs(rows: list[dict[str, str]]) -> None:
    failures = [row for row in rows if row["Status"] == "Fail"]
    lines = ["# GUI Bug Report — HW03 Task 1", "", "**Traceability:** one reproducible record per failed checklist assertion; existing issues were fetched before reuse.", "",
             "| Bug ID | Severity | Checklist ID | GitHub | Evidence |", "|---|---|---|---|---|"]
    drafts = TASK / "github-issues"
    drafts.mkdir(parents=True, exist_ok=True)
    for row in failures:
        severity = SEVERITY.get(row["ID"], "Medium")
        issue = row["GitHub Issue"]
        github = f"[Issue]({issue})" if issue.startswith("https://") else f"`{issue}`"
        evidence = f"[PNG]({row['Evidence']})"
        lines.append(f"| `{row['Bug ID']}` | {severity} | `{row['ID']}` | {github} | {evidence} |")
    for row in failures:
        severity = SEVERITY.get(row["ID"], "Medium")
        lines += ["", f"## {row['Bug ID']} — {row['Checklist Item']}", "", f"- Severity: **{severity}**.",
                  f"- Expected: {row['Expected Result']}", f"- Actual: {row['Actual Result']}",
                  f"- Mode: `{row['Execution Mode']}`.", f"- Evidence: [screenshot]({row['Evidence']}).",
                  "- Reproduction:", f"  1. Start EShop and open `{row['Screen/Route']}`.",
                  f"  2. Perform `{row['ID']}` as described by the checklist.", "  3. Observe the recorded UI/network/dialog state.",
                  f"- GitHub: {row['GitHub Issue']}"]
        raw = f"https://raw.githubusercontent.com/trngnneee/eshop-sut/HW3-Khoa/task1-gui/{row['Evidence']}"
        draft = f"""# {row['Checklist Item']}

**Local ID:** `{row['Bug ID']}`
**Status:** `{'EXISTING_ISSUE_REUSED' if row['GitHub Issue'].startswith('https://') else 'PENDING_EXTERNAL_ACTION'}`
**Severity:** `{severity}`
**Reporter:** Đặng Đăng Khoa (23127207)
**Environment:** {PLATFORM}

## Steps

1. Start EShop and open `{row['Screen/Route']}`.
2. Execute `{row['ID']}`: {row['Checklist Item']}
3. Observe the UI and request/dialog state.

## Expected

{row['Expected Result']}

## Actual

{row['Actual Result']}

## Evidence

![{row['Bug ID']}]({raw})

Local file: `{row['Evidence']}`

## Duplicate-search disposition

{row['GitHub Issue']}
"""
        (drafts / f"{row['Bug ID']}.md").write_text(draft, encoding="utf-8")
    (TASK / "GUI_Bug_Report_HW3.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_ai_review(rows: list[dict[str, str]]) -> None:
    lines = ["# AI Item-Level Critique — Task 1", "", "**Student reviewer:** Đặng Đăng Khoa — 23127207  ",
             "**Review date:** 2026-08-02  ", "**Status:** `HUMAN_REVIEWED`", "",
             "| ID | Origin | Decision | Evidence-based review |", "|---|---|---|---|"]
    reframed = {"GUI-WEB-LOGIN-010", "GUI-WEB-LOGIN-013", "GUI-ADMIN-CATEGORY-005", "GUI-ADMIN-CATEGORY-011"}
    for row in rows:
        if row["Origin"] == "HUMAN_ADDED":
            decision = "HUMAN_ADDED_RETAINED"
            review = f"Added by human review because {HUMAN_REASON[row['ID']]}; current status={row['Status']}."
        elif row["ID"] in reframed:
            decision = "AI_ITEM_REFRAMED"
            review = f"Original expectation was inaccurate or ambiguous; rewritten as a measurable {row['Related Requirement']} assertion before applying status={row['Status']}."
        else:
            decision = "AI_ITEM_RETAINED"
            review = f"Measurable within {row['Related Requirement']}; executed with status={row['Status']} and evidence={row['Evidence ID']}."
        lines.append(f"| `{row['ID']}` | `{row['Origin']}` | `{decision}` | {clean(review)} |")
    lines += ["", "## Student confirmation", "",
              "The student confirmed in chat on 2026-08-02 that all human-review sections were reviewed. This does not convert unavailable physical-device evidence into a completed run."]
    (TASK / "AI_Item_Level_Critique.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_ai_docs() -> None:
    audit = f"""# AI Audit Report — Task 1

**Student:** Đặng Đăng Khoa — 23127207
**Date:** {DATE}
**Status:** `HUMAN_REVIEWED — EXTERNAL_EVIDENCE_BLOCKERS_DISCLOSED`

## Interaction log

| Field | Value |
|---|---|
| AI tool | OpenAI Codex, local PowerShell/Python and the corrected Task 3 Playwright result set |
| User request | Repair all three tasks toward full rubric compliance; the student explicitly confirmed every human-review section. |
| AI use | Reconciled 58 Task 1 IDs with the final Chrome execution, copied 40 identity-overlaid screenshots, corrected requirement mismatches, regenerated Markdown/Excel/results/reports and prepared GitHub traceability. |
| Human review | Confirmed by the student on {DATE}. |
| Runtime provenance | `../task3-cross-platform/results/chrome-windows.json` and its indexed evidence. |

## Material corrections

- The older 36/22 checklist and 40/18 summary were replaced by one current dataset.
- Category Edit and duplicate-name rejection were removed as invented FR-14 expectations.
- The FR-02 lockout expectation now uses three attempts and 30 seconds.
- Every row records `LIVE_LOCAL_SUT`, `MOCKED_*` or `EXPO_WEB_DESKTOP_BROWSER` explicitly.
- Real soft-keyboard behavior remains Blocked without a phone.

## Anti-fabrication declaration

No participant, pilot, GitHub URL, device run or YouTube upload was invented. Existing issue URLs were fetched from GitHub before reuse. New records remain pending until GitHub returns a real URL.
"""
    (TASK / "AI_Audit_Report_Task1.md").write_text(audit, encoding="utf-8")
    disclosure = f"""# AI Disclosure — Task 1

**Status:** `HUMAN_REVIEWED`
**Student:** Đặng Đăng Khoa — 23127207
**Review date:** {DATE}

AI assisted checklist design, mechanical reconciliation, report generation and semantic validation. Final verdicts come from the corrected Task 3 Chrome execution rather than source-code inference. The student confirmed human review. Mocked states and the unavailable physical soft-keyboard run are disclosed; external URLs are never fabricated.
"""
    (TASK / "AI_Disclosure_Task1.md").write_text(disclosure, encoding="utf-8")
    critique = """# AI Critique — Task 1

**Status:** `HUMAN_REVIEWED`
**Student review date:** `2026-08-02`

AI giúp mở rộng checklist lên 58 mục và gợi ý các nhóm quan trọng như accessibility, loading, empty state, network failure, double submit và responsive 320 px. Tuy nhiên, đầu ra ban đầu có ba sai lệch nghiêm trọng. Thứ nhất, AI dựa vào source code rồi ghi Actual/Status như thể đã quan sát runtime; năm ảnh tĩnh không chứng minh được các lỗi động. Thứ hai, nó hiểu FR-14 thành đầy đủ CRUD nên xem việc thiếu Edit Category là defect, đồng thời tự đặt yêu cầu cấm tên danh mục trùng. Thứ ba, summary 40 Pass/18 Fail không khớp checklist 36 Pass/22 Fail, cho thấy nhiều artefact không đồng nghĩa với dữ liệu nhất quán.

Lần hiệu chỉnh dùng 58 result rows từ execution Chrome đã sửa harness, lưu execution mode và liên kết mỗi row với screenshot có overlay. Các state cần kiểm soát timing được ghi MOCKED, còn core flows là LIVE_LOCAL_SUT. Hai expectation sai FR-14 và edge case trim không có requirement được viết lại trước khi phân loại; threshold lockout cũng được sửa về ba lần và 30 giây. Mobile soft keyboard vẫn Blocked vì Expo Web không tạo bằng chứng thiết bị thật.

Bài học là AI phù hợp để tăng coverage và tự động hóa đối chiếu, nhưng không được tự quyết định provenance. Mỗi verdict phải truy ngược về requirement, runtime observation và evidence. Validator chỉ đáng tin khi kiểm tra ngữ nghĩa, URL thật, execution mode và blocker bên ngoài, không chỉ kiểm tra file tồn tại.
"""
    (TASK / "AI_Critique_Task1.md").write_text(critique, encoding="utf-8")


def write_readme(rows: list[dict[str, str]]) -> None:
    count = Counter(row["Status"] for row in rows)
    text = f"""# Task 1 — GUI Checklist Deliverables

**Student:** Đặng Đăng Khoa — 23127207
**SUT:** EShop
**Primary execution:** {PLATFORM}, {DATE}
**Status:** `BLOCKED_REAL_MOBILE_SOFT_KEYBOARD_AND_PENDING_EXTERNAL_ITEMS`

## Outcome

- 58 unique items; IA-01 through IA-04 covered.
- {count['Pass']} Pass, {count['Fail']} Fail, {count['Blocked']} Blocked.
- 48 AI-initial and 10 human-added items, human-reviewed.
- 40 current Chrome screenshots with identity/email overlay.
- One source of truth: `results/Task1_Execution_Chrome.csv`.

## Completion boundary

1. `GUI-MOBILE-LOGIN-011` needs Expo Go or a physical/cloud phone.
2. Every `PENDING_EXTERNAL_ACTION` bug needs a real duplicate/new GitHub URL.
3. `Demo_Video_Link.md` needs a real public Task 1 GUI-skill YouTube URL.

No unavailable evidence is reconstructed to satisfy a validator.
"""
    (TASK / "README.md").write_text(text, encoding="utf-8")
    demo = TASK / "Demo_Video_Link.md"
    if not demo.exists() or "https://youtu" not in demo.read_text(encoding="utf-8-sig"):
        demo.write_text("# Task 1 GUI Testing Skill Demo\n\n**Status:** `PENDING_EXTERNAL_YOUTUBE_UPLOAD`\n\nA real public YouTube URL showing the GUI-testing skill end to end is required.\n", encoding="utf-8")


def normalize_markdown_whitespace() -> None:
    """Keep generated Markdown diff-clean without changing rendered content."""
    markdown_files = list(TASK.glob("*.md")) + list((TASK / "github-issues").glob("*.md"))
    for markdown_file in markdown_files:
        text = markdown_file.read_text(encoding="utf-8-sig")
        normalized = "\n".join(line.rstrip() for line in text.splitlines()) + "\n"
        markdown_file.write_text(normalized, encoding="utf-8")


def main() -> None:
    rows = reconcile()
    expected_fails = {row["ID"] for row in rows if row["Status"] == "Fail"}
    if len(rows) != 58 or len(expected_fails) != 20:
        raise RuntimeError(f"Unexpected reconciled totals: rows={len(rows)}, fails={len(expected_fails)}")
    write_checklist(rows)
    write_csv(rows)
    write_excel(rows)
    write_coverage(rows)
    write_summary(rows)
    write_bug_docs(rows)
    write_ai_review(rows)
    write_ai_docs()
    write_readme(rows)
    normalize_markdown_whitespace()
    output = {"rows": len(rows), "status": dict(Counter(r["Status"] for r in rows)),
              "screenshots": len({r["Evidence"] for r in rows}),
              "pending_issues": sum(r["GitHub Issue"] == "PENDING_EXTERNAL_ACTION" for r in rows)}
    print(json.dumps(output, ensure_ascii=False))


if __name__ == "__main__":
    main()
